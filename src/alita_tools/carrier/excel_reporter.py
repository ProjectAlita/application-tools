import pandas as pd

import time
from datetime import datetime
import numpy as np
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import (Alignment, Font, PatternFill, Border, Side)
from openpyxl.formatting.rule import CellIsRule

from typing import Dict, Any

RED_COLOR = 'F7A9A9'
GREEN_COLOR = 'AFF2C9'
YELLOW_COLOR = 'F7F7A9'
RED_COLOR_FONT = '00F90808'
GREEN_COLOR_FONT = '002BBD4D'
RED_FILL = PatternFill(start_color=RED_COLOR, end_color=RED_COLOR, fill_type='solid')
GREEN_FILL = PatternFill(start_color=GREEN_COLOR, end_color=GREEN_COLOR, fill_type='solid')
YELLOW_FILL = PatternFill(start_color=YELLOW_COLOR, end_color=YELLOW_COLOR, fill_type='solid')


class PerformanceReportParser:

    def parse(self):
        raise NotImplementedError("Subclasses should implement this method")


class JMeterReportParser(PerformanceReportParser):

    def __init__(self, file_path, calculated_think_time):
        self.file_path = file_path
        self.calculated_think_time = calculated_think_time

    def parse(self):
        transactions_data = dict(requests={})
        pd_data = pd.read_csv(self.file_path, sep=',')
        if pd_data.responseMessage.isnull().all():
            data_wo_transactions = pd_data
        else:
            data_wo_transactions = pd_data[
                ~pd_data.responseMessage.str.contains(r'^Number of samples', case=False, na=False)]
        summary_ = self.calculate_statistics(data_wo_transactions, 'Total')

        transactions = pd_data.label.unique()
        for transaction in transactions:
            data_ = pd_data[pd_data['label'] == transaction]
            transactions_data['requests'].update({transaction: self.calculate_statistics(data_, transaction)})
        transactions_data['requests'].update(
            {'Total Transactions': self.calculate_statistics(pd_data, 'Total Transactions')})
        transactions_data.update({
            'max_user_count': summary_['max_user_count'],
            'ramp_up_period': summary_['ramp_up_period'],
            'error_rate': summary_['error_rate'],
            'date_start': summary_['date_start'],
            'date_end': summary_['date_end'],
            'throughput': summary_['throughput'],
            'duration': summary_['duration'],
            'think_time': self.calculated_think_time
        })
        return transactions_data

    def convert_time(self, ts):
        element = datetime.strptime(ts, "%Y-%m-%d %H:%M:%S")
        tuple = element.timetuple()
        timestamp = time.mktime(tuple)
        return timestamp

    def calculate_statistics(self, df, req_name):
        df = df.sort_values(by=['timeStamp'])
        total_error_count = len(df.loc[df['success'] != True])
        total_request = df['elapsed'].size
        ok_req = total_request - total_error_count
        summary = {
            'request_name': req_name,
            'min': round(float(df['elapsed'].min()), 3),
            'max': round(float(df['elapsed'].max()), 3),
            'Total': total_request,
            'average': round(float(df['elapsed'].mean()), 3),
            'median': round(float(df['elapsed'].median()), 3),
            '90Pct': round(df['elapsed'].quantile(0.9), 3),
            '95Pct': round(df['elapsed'].quantile(0.95), 3),
            'KO': total_error_count,
            'OK': ok_req,
            'Error%': round(total_error_count / total_request, 4)
        }
        if req_name == 'Total':
            end_test = df.timeStamp.iloc[-1]
            start_test = df.timeStamp.iloc[0]

            try:
                duration = pd.to_timedelta(end_test - start_test, unit='ms')
            except TypeError as ex:
                print(ex)
                end_test = self.convert_time(end_test)
                start_test = self.convert_time(start_test)
                duration = pd.to_timedelta(end_test - start_test, unit='ms')
            throughput = round(ok_req / duration.seconds, 2)
            duration_ = str(duration).split('days')[1].split('.')[0]
            ramp_up_end = df.loc[df['allThreads'] == df.allThreads.max()].timeStamp.min()
            ramp_up = str(pd.to_timedelta(ramp_up_end - df.timeStamp.iloc[0], unit='ms')).split('days')[1].split('.')[0]
            summary['duration'] = duration_
            summary['ramp_up_period'] = ramp_up
            summary['throughput'] = throughput
            summary['error_rate'] = round(total_error_count / total_request, 4) * 100
            unique_thread_arr = df.threadName.unique()
            filter_unique_thread_arr = list(filter(lambda x: not x.startswith('parallel'), unique_thread_arr))
            summary['max_user_count'] = len(filter_unique_thread_arr)
            summary['date_start'] = datetime.fromtimestamp(int(start_test / 1000)).strftime(
                '%Y-%m-%d %H:%M:%S')
            summary['date_end'] = datetime.fromtimestamp(int(end_test / 1000)).strftime(
                '%Y-%m-%d %H:%M:%S')
        # else: why do we need it ?
        #     summary['total_response_time'] = list(df.elapsed)
        return summary


class GatlingReportParser(PerformanceReportParser):

    def __init__(self, log_file: str, think_times="5,0-10,0"):
        self.calculated_think_time = think_times
        self.log_file = log_file

    @staticmethod
    def convert_timestamp_to_datetime(timestamp: int) -> datetime:
        """Converts a timestamp to a human-readable datetime object."""
        return datetime.fromtimestamp(timestamp / 1000.0)

    def parse(self) -> Dict[str, Any]:
        latest_log_file = self.log_file
        print(f"Path: {latest_log_file}")
        groups, requests, users, date_start, date_end, ramp_up_period = self.parse_log_file(latest_log_file)

        # Combine requests and groups for reporting purposes
        requests = defaultdict(list, {**requests})
        duration = self.calculate_duration(date_start, date_end)

        transactions_data = {
            "requests": {
                transaction: self.calculate_single_metric(transaction, entries)
                for transaction, entries in requests.items()
            },
            "groups": {
                group: self.calculate_single_metric(group, entries)
                for group, entries in groups.items()
            }
        }
        # Calculate total requests only from the original requests dictionary
        total_requests = self.calculate_all_requests(requests)
        transactions_data["requests"]["Total Requests"] = total_requests

        throughput = total_requests['Total'] / (duration * 60)  # Initializing a variable with the default value

        if duration > 1:
            duration = round(duration)
        if ramp_up_period > 1:
            ramp_up_period = round(ramp_up_period, 1)
            throughput = total_requests['Total'] / (duration * 60)
        if throughput > 1:
            throughput = round(throughput, 1)

        # Add any additional required summary metrics
        # Convert to Eastern Time (EST/EDT)
        from pytz import timezone
        eastern = timezone('US/Eastern')
        date_start = date_start.astimezone(eastern) if date_start else None
        date_end = date_end.astimezone(eastern) if date_end else None

        transactions_data.update({
            'max_user_count': users,
            'ramp_up_period': ramp_up_period,
            'error_rate': total_requests['Error%'],
            'date_start': date_start.strftime('%Y-%m-%d %H:%M:%S') if date_start else None,
            'date_end': date_end.strftime('%Y-%m-%d %H:%M:%S') if date_end else None,
            'throughput': throughput,
            'duration': duration,
            'think_time': self.calculated_think_time
        })
        return transactions_data

    def parse_log_file(self, file_path: str):
        groups = defaultdict(list)
        requests = defaultdict(list)
        users = 0
        date_start = None
        date_end = None
        ramp_start = None
        ramp_end = None
        try:
            with open(file_path, 'r', encoding="utf8") as file:
                for line_number, line in enumerate(file):
                    if not date_start:
                        if 'ASSERTION' in line:
                            continue
                        date_start = self.convert_timestamp_to_datetime(int(line.split('\t')[3]))
                    try:
                        date_end = self.convert_timestamp_to_datetime(int(line.split('\t')[3]))
                    except:
                        # TODO: Fix logic
                        pass
                    if line.startswith('REQUEST'):
                        self.parse_request_line(requests, line)
                    elif line.startswith('USER') and 'START' in line:
                        users += 1
                        if "START" in line:
                            if not ramp_start:
                                ramp_start = self.convert_timestamp_to_datetime(int(line.split('\t')[3]))
                            else:
                                ramp_end = self.convert_timestamp_to_datetime(int(line.split('\t')[3]))

                    elif line.startswith('GROUP'):
                        self.parse_group_line(groups, line)
        except FileNotFoundError as e:
            print(f"File not found: {e}")
            raise
        except Exception as e:
            print(f"Error reading file {file_path}: {e}")
            raise
        return groups, requests, users, date_start, date_end, self.calculate_duration(ramp_start, ramp_end)

    @staticmethod
    def calculate_duration(date_start: datetime, date_end: datetime) -> float:
        """
        Calculate the total duration between the first and last timestamp in the logs.
        """
        if not all([date_start, date_end]):
            return 0.0
        test_dur = round((date_end - date_start).total_seconds() / 60, 3)
        print(f"Test duration, min: {test_dur}")
        return test_dur

    @staticmethod
    def parse_request_line(requests, line):
        parts = line.split('\t')
        if len(parts) >= 7:
            request_name = parts[2]
            start_time = int(parts[3])
            end_time = int(parts[4])
            status = parts[5].strip()
            response_time = (end_time - start_time)
            requests[request_name].append((response_time, status))

    @staticmethod
    def parse_group_line(groups, line):
        parts = line.split('\t')
        if len(parts) >= 6:
            group_name = parts[1]
            response_time = int(parts[4])
            status = parts[5].strip()
            groups[group_name].append((response_time, status))

    def calculate_single_metric(self, name, entries):
        response_times = [d[0] for d in entries]
        ko_count = len([d for d in entries if d[1] != 'OK'])
        total_count = len(entries)
        ko_percentage = round((ko_count / total_count), 4) if total_count > 0 else 0

        if response_times:
            min_time, avg_time, p50_time, p90_time, p95_time, max_time = self.calculate_statistics(response_times)
        else:
            min_time = avg_time = p50_time = p90_time = p95_time = max_time = 0

        return {
            "request_name": name,
            "Total": total_count,
            "KO": ko_count,
            "Error%": ko_percentage,
            "min": min_time,
            "average": avg_time,
            "90Pct": p90_time,
            "95Pct": p95_time,
            "max": max_time
        }

    @staticmethod
    def calculate_statistics(response_times):
        min_time = round(min(response_times), 3)
        avg_time = round(np.mean(response_times), 3)
        p50_time = round(np.percentile(response_times, 50), 3)  # Median
        p90_time = round(np.percentile(response_times, 90), 3)
        p95_time = round(np.percentile(response_times, 95), 3)
        max_time = round(max(response_times), 3)
        return min_time, avg_time, p50_time, p90_time, p95_time, max_time

    def calculate_all_requests(self, requests: defaultdict) -> dict:
        all_entries = [entry for entries in requests.values() for entry in entries]
        return self.calculate_single_metric('Total', all_entries)


class ExcelReporter(object):

    def __init__(self, report_path="/tmp/lr_results.xlsx"):
        self.report_path = report_path
        self.header = {
            'Transaction': 'request_name',
            'Req, count': 'Total',
            'KO, count': 'KO',
            'KO, %': 'Error%',
            'Min, sec': 'min',
            'Avg, sec': 'average',
            '90p, sec': '90Pct',
            '95p, sec': '95Pct',
            'Max, sec': 'max',
        }
        self.title = {}

    def prepare_headers_and_titles(self):
        self.title = {
            'Users': 'max_user_count',
            'Ramp Up, min': 'ramp_up_period',
            'Duration, min': 'duration',
            'Think time, sec': 'think_time',
            'Start Date, EST': 'date_start',
            'End Date, EST': 'date_end',
            'Throughput, req/sec': 'throughput',
            'Error rate, %': 'error_rate',
            'Carrier report': 'carrier_report',
            'Build status': 'build_status',
            "Justification": 'justification'
        }

    def write_to_excel(self, results, carrier_report, thresholds, report_percentile):
        results["carrier_report"] = carrier_report
        try:
            results["build_status"], results["justification"] = self.get_build_status_and_justification(thresholds,
                                                                                                    report_percentile)
        except Exception as e:
            print(e)
            results["build_status"], results["justification"] = "SUCCESS", ""
        wb = Workbook()
        ws = wb.active
        ws.title = "Test results"

        requests_filtered = results["requests"]

        self.write_title_above_headers(self.title, results, self.header, ws)

        content_start, content_end = None, None
        for i, (header_field_name, header_key) in enumerate(self.header.items()):
            header_cell = ws.cell(row=len(self.title) + 1, column=i + 1,
                                  value=header_field_name)  # i + 1 because excel index starts from 1
            header_cell.alignment = Alignment(horizontal="center")
            border_style = Side(border_style="thin", color="040404")
            header_cell.border = Border(top=border_style, left=border_style, right=border_style, bottom=border_style)
            header_cell.fill = PatternFill("solid", fgColor='007FD5D8')

            if not content_start:
                content_start = get_column_letter(header_cell.column) + str(header_cell.row)

            need_cond_format = header_key in ['min', 'average', '95Pct', 'max', '90Pct']
            start_cell_for_formatting, end_cell_for_formatting = None, None
            for j, req_info in enumerate(requests_filtered.values()):
                if need_cond_format:
                    cur_cell = ws.cell(row=j + len(self.title) + 2, column=i + 1,
                                       value=round(req_info[header_key] / 1000, 3))
                    if start_cell_for_formatting is None:
                        start_cell_for_formatting = cur_cell
                    end_cell_for_formatting = cur_cell
                else:
                    cur_cell = ws.cell(row=j + len(self.title) + 2, column=i + 1, value=req_info[header_key])
                border_style = Side(border_style="thin", color="040404")
                cur_cell.border = Border(top=border_style, left=border_style, right=border_style, bottom=border_style)
                if header_key == 'Error%':
                    cur_cell.number_format = '0.00%'
                content_end = get_column_letter(cur_cell.column) + str(cur_cell.row)

            if need_cond_format:
                start_col = get_column_letter(start_cell_for_formatting.column)
                start_row = start_cell_for_formatting.row
                end_col = get_column_letter(end_cell_for_formatting.column)
                end_row = end_cell_for_formatting.row
                coords = start_col + str(start_row) + ':' + end_col + str(end_row)
                try:
                    rt_threshold = float(self.get_response_threshold(thresholds))
                except Exception as e:
                    print(e)
                    rt_threshold = 0
                if rt_threshold > 1:
                    threshold = rt_threshold/1000
                else:
                    threshold = rt_threshold

                ws.conditional_formatting.add(coords,
                                              CellIsRule(operator='lessThan', formula=[threshold], fill=GREEN_FILL))
                ws.conditional_formatting.add(coords, CellIsRule(operator='greaterThan', formula=[threshold * 1.5],
                                                                 fill=RED_FILL))
                ws.conditional_formatting.add(coords,
                                              CellIsRule(operator='between', formula=[threshold, threshold * 1.5],
                                                         fill=YELLOW_FILL))

                # column width adjusted
        for i, key in enumerate(self.header):
            length = len(str(key))
            ws.column_dimensions[get_column_letter(i + 1)].width = length + 5

        first_column_length = max(len(str(cell.value)) for cell in ws["A"])
        ws.column_dimensions[get_column_letter(1)].width = first_column_length + 5

        ws.auto_filter.ref = content_start + ":" + content_end

        # horizontal and vertical freeze
        freezen_area = ws['B' + str(len(self.title) + 2)]
        ws.freeze_panes = freezen_area
        print(f"Excel path: {self.report_path}")
        wb.save(self.report_path)

    def get_build_status_and_justification(self, thresholds, report_percentile):
        justification = ""
        build_status = "SUCCESS"
        rt_threshold = self.get_response_threshold(thresholds)
        for th in thresholds:
            if th["target"] == "error_rate":
                if th["status"] == "FAILED":
                    justification = "Total error rate exceed the threshold of " + str(th["threshold"]) + "% "
                    build_status = "FAILED"
                else:
                    justification = "Total error rate doesn't exceed the threshold of " + str(th["threshold"]) + "% "
            if th["target"] == "response_time":
                if th["status"] == "FAILED":
                    if build_status == "FAILED":
                        justification += "and Response Time for some transaction(s) exceed the threshold of " + str(
                            rt_threshold / 1000) + " seconds by " + report_percentile
                    else:
                        build_status = "FAILED"
                        justification = "Response Time for some transaction(s) exceed the threshold of " + str(
                            rt_threshold / 1000) + " seconds by " + report_percentile
                    return build_status, justification
        justification += "and Response Time for all transactions doesn't exceed the threshold of " + str(
            rt_threshold / 1000) + " seconds by " + report_percentile
        return build_status, justification

    @staticmethod
    def get_response_threshold(thresholds):
        for th in thresholds:
            if th["target"] == "response_time":
                return th["threshold"]

    @staticmethod
    def write_title_above_headers(title, results, header, ws):

        for i, (header_field_name, header_key) in enumerate(title.items()):
            title_name = ws.cell(row=i + 1, column=1, value=header_field_name)
            title_value = ws.cell(row=i + 1, column=2, value=results[header_key])
            title_name.font = Font(b=True, color='00291A75')
            title_name.fill = PatternFill("solid", fgColor='00CDEBEA')
            title_name.alignment = Alignment(horizontal="left", vertical="center")
            title_value.alignment = Alignment(horizontal="center", vertical="center")
            title_value.fill = PatternFill("solid", fgColor='00CDEBEA')
            border_style = Side(border_style="thin", color="040404")
            title_value.border = Border(top=border_style, left=border_style, right=border_style, bottom=border_style)
            title_name.border = Border(top=border_style, left=border_style, right=border_style, bottom=border_style)

            if header_key == 'error_rate':
                title_value.number_format = '0.00%'
            if header_key == 'carrier_report':
                title_value.hyperlink, title_value.value = title_value.value, "Carrier report"
                title_value.font = Font(b=True, underline="single", color='00291A75')
            if header_key == 'build_status':
                title_value.font = Font(b=True,
                                        color=GREEN_COLOR_FONT if title_value.value == 'SUCCESS'
                                        else RED_COLOR_FONT)
            if header_key == 'justification' and len(title_value.value) > 125:
                title_value.alignment = Alignment(horizontal="center", vertical="justify")
                ws.row_dimensions[i + 1].height = 30

            ws.merge_cells(start_row=i + 1, start_column=2, end_row=i + 1, end_column=len(header.items()))
